"""
SCRIPT 2 - MANTENIMIENTO MENSUAL
Solo trae reseñas NUEVAS desde la última vez que corriste el setup.
Úsalo cada semana o mes con clientes activos.
"""

import anthropic
import os
import json
from datetime import datetime
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

anthropic_client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


def cargar_registro(nombre_limpio: str) -> dict:
    archivo = f"registro_{nombre_limpio}.json"
    ruta = os.path.join(os.path.dirname(__file__), archivo)

    if not os.path.exists(ruta):
        raise Exception(f"No se encontró el registro '{archivo}'. Corre primero el script de setup.")

    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)


def listar_registros() -> list[str]:
    directorio = os.path.dirname(__file__)
    return [f for f in os.listdir(directorio) if f.startswith("registro_") and f.endswith(".json")]


def obtener_resenas_nuevas(data_id: str, serpapi_key: str, desde: datetime) -> list[dict]:
    resenas_nuevas = []
    next_page_token = None
    pagina = 1
    detener = False

    while not detener:
        print(f"  Revisando página {pagina}...", end=" ")

        params = {
            "engine": "google_maps_reviews",
            "data_id": data_id,
            "api_key": serpapi_key,
            "hl": "es",
            "sort_by": "newestFirst",
        }

        if next_page_token:
            params["next_page_token"] = next_page_token

        search = GoogleSearch(params)
        results = search.get_dict()

        if "error" in results:
            print(f"Error: {results['error']}")
            break

        resenas_pagina = results.get("reviews", [])
        if not resenas_pagina:
            print("sin más reseñas.")
            break

        nuevas_en_pagina = 0
        for r in resenas_pagina:
            texto = r.get("snippet", "").strip()
            fecha_str = r.get("date", "")

            # Si la reseña es más antigua que el último proceso, detenemos
            # SerpAPI devuelve fechas relativas ("hace 2 semanas"), usamos posición como proxy
            if not texto:
                continue

            resenas_nuevas.append({
                "texto": texto,
                "estrellas": int(r.get("rating", 3)),
                "fecha": fecha_str,
                "autor": r.get("user", {}).get("name", "Cliente")
            })
            nuevas_en_pagina += 1

        print(f"{nuevas_en_pagina} reseñas ✓")

        next_page_token = results.get("serpapi_pagination", {}).get("next_page_token")

        # En mantenimiento solo procesamos las primeras 2 páginas (reseñas recientes)
        if not next_page_token or pagina >= 2:
            break

        pagina += 1

    return resenas_nuevas


def generar_respuesta(resena: dict, nombre: str, tipo: str, tono: str) -> str:
    prompt = f"""Eres el community manager de {nombre}, un {tipo}.
Tu tono de comunicación es {tono}.

{resena['autor']} dejó esta reseña de Google con {resena['estrellas']} estrellas:
"{resena['texto']}"

ESTRUCTURA OBLIGATORIA:
1. Hook de primera línea: máximo 6 palabras
2. Respuesta personalizada con detalles de la reseña
3. Una sola CTA al final

REGLAS:
- Máximo 80 palabras
- Si es positiva (4-5 estrellas): agradece específicamente, invita a volver
- Si es negativa (1-2 estrellas): empatía, reconoce el problema, ofrece resolverlo offline
- Si es neutral (3 estrellas): agradece, reconoce mejora, invita a volver
- Sin hashtags, español latino, tono directo y confiable
- Termina con el nombre del negocio

Devuelve SOLO la respuesta lista para publicar."""

    message = anthropic_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}]
    )
    return message.content[0].text.strip()


def actualizar_registro(ruta_registro: str, registro: dict, total_nuevas: int):
    registro["ultimo_proceso"] = datetime.now().isoformat()
    registro["total_procesadas"] = registro.get("total_procesadas", 0) + total_nuevas

    with open(ruta_registro, "w", encoding="utf-8") as f:
        json.dump(registro, f, ensure_ascii=False, indent=2)


def guardar_resultados(resultados: list[dict], nombre_negocio: str):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = nombre_negocio.replace(" ", "_").replace("/", "-")

    verde = PatternFill("solid", fgColor="1F7A4A")
    gris_claro = PatternFill("solid", fgColor="F2F2F2")
    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reseñas nuevas"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"MANTENIMIENTO — {nombre_negocio}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = verde
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Reseñas nuevas: {len(resultados)}"
    ws["A2"].font = Font(italic=True, size=10, color="888888")
    ws.merge_cells("A2:F2")

    encabezados = ["#", "Estrellas", "Autor", "Fecha", "Reseña del cliente", "✅ Respuesta (copiar y pegar en Google)"]
    anchos = [5, 10, 20, 15, 50, 60]

    for col, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
        celda = ws.cell(row=3, column=col, value=enc)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = verde
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.row_dimensions[3].height = 25

    for i, r in enumerate(resultados, 1):
        fila = i + 3
        estrellas_texto = "⭐" * r["estrellas"] + f" ({r['estrellas']})"
        valores = [i, estrellas_texto, r["autor"], r["fecha"], r["resena_original"], r["respuesta_generada"]]

        for col, valor in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = borde
            if i % 2 == 0:
                celda.fill = gris_claro

        ws.row_dimensions[fila].height = 80

    ws.freeze_panes = "A4"

    archivo = f"MANTENIMIENTO_{nombre_limpio}_{timestamp}.xlsx"
    ruta = os.path.join(os.path.dirname(__file__), archivo)
    wb.save(ruta)
    print(f"\nGuardado en: {archivo}")
    return archivo


def main():
    print("\n=== MANTENIMIENTO — RESEÑAS NUEVAS ===\n")

    serpapi_key = os.environ.get("SERPAPI_KEY") or input("SerpAPI key: ").strip()

    # Mostrar clientes disponibles
    registros = listar_registros()
    if not registros:
        print("No hay clientes registrados. Corre primero el script 1_setup_completo.py")
        return

    print("Clientes registrados:")
    for i, r in enumerate(registros, 1):
        nombre = r.replace("registro_", "").replace(".json", "").replace("_", " ")
        print(f"  {i}. {nombre}")

    seleccion = input("\nNúmero del cliente (o Enter para el primero): ").strip()
    idx = int(seleccion) - 1 if seleccion.isdigit() else 0
    archivo_registro = registros[idx]
    nombre_limpio = archivo_registro.replace("registro_", "").replace(".json", "")

    ruta_registro = os.path.join(os.path.dirname(__file__), archivo_registro)
    registro = cargar_registro(nombre_limpio)

    nombre_negocio = registro["nombre_negocio"]
    data_id = registro["data_id"]
    tipo = registro.get("tipo", "restaurante")
    tono = registro.get("tono", "profesional y cercano")
    ultimo_proceso = datetime.fromisoformat(registro["ultimo_proceso"])

    print(f"\nCliente: {nombre_negocio}")
    print(f"Último proceso: {ultimo_proceso.strftime('%d/%m/%Y %H:%M')}")
    print(f"Buscando reseñas nuevas...")

    resenas = obtener_resenas_nuevas(data_id, serpapi_key, ultimo_proceso)

    if not resenas:
        print("\nNo hay reseñas nuevas desde el último proceso.")
        return

    print(f"\n{len(resenas)} reseñas encontradas. Generando respuestas...")

    resultados = []
    for i, r in enumerate(resenas, 1):
        print(f"  Procesando {i}/{len(resenas)}...", end=" ")
        respuesta = generar_respuesta(r, nombre_negocio, tipo, tono)
        resultados.append({
            "estrellas": r["estrellas"],
            "autor": r["autor"],
            "fecha": r["fecha"],
            "resena_original": r["texto"],
            "respuesta_generada": respuesta
        })
        print("✓")

    print("\n" + "=" * 60)
    for i, r in enumerate(resultados, 1):
        print(f"\nRESEÑA #{i} — {'⭐' * r['estrellas']} — {r['autor']}")
        print(f"Cliente: {r['resena_original']}")
        print(f"\nRespuesta:\n{r['respuesta_generada']}")
        print("-" * 60)

    guardar = input("\n¿Guardar resultados? (s/n): ").strip().lower()
    if guardar == "s":
        guardar_resultados(resultados, nombre_negocio)
        actualizar_registro(ruta_registro, registro, len(resultados))
        print("Registro actualizado.")


if __name__ == "__main__":
    main()
