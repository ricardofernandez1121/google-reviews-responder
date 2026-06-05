"""
SCRIPT 1 - SETUP COMPLETO
Trae TODAS las reseñas históricas del negocio y genera respuestas.
Úsalo la primera vez con un cliente nuevo.
"""

import anthropic
import os
import json
from datetime import datetime
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

anthropic_client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


def buscar_data_id(nombre_negocio: str, serpapi_key: str) -> tuple[str, str]:
    params = {
        "engine": "google_maps",
        "q": nombre_negocio,
        "api_key": serpapi_key,
        "hl": "es",
    }
    search = GoogleSearch(params)
    results = search.get_dict()

    lugares = results.get("local_results", [])
    if not lugares:
        raise Exception("No se encontró el negocio. Intenta con un nombre más específico.")

    lugar = lugares[0]
    print(f"Negocio encontrado: {lugar.get('title')}")
    return lugar.get("data_id", ""), lugar.get("title", nombre_negocio)


def obtener_todas_las_resenas(data_id: str, serpapi_key: str) -> list[dict]:
    resenas = []
    next_page_token = None
    pagina = 1

    while True:
        print(f"  Descargando página {pagina}...", end=" ")

        params = {
            "engine": "google_maps_reviews",
            "data_id": data_id,
            "api_key": serpapi_key,
            "hl": "es",
        }

        if next_page_token:
            params["next_page_token"] = next_page_token

        search = GoogleSearch(params)
        results = search.get_dict()

        if "error" in results:
            print(f"Error: {results['error']}")
            break

        resenas_pagina = results.get("reviews", [])
        if not resenas_pagina:
            print("sin más reseñas.")
            break

        for r in resenas_pagina:
            texto = r.get("snippet", "").strip()
            if texto:
                resenas.append({
                    "texto": texto,
                    "estrellas": int(r.get("rating", 3)),
                    "fecha": r.get("date", ""),
                    "autor": r.get("user", {}).get("name", "Cliente")
                })

        print(f"{len(resenas_pagina)} reseñas ✓")

        next_page_token = results.get("serpapi_pagination", {}).get("next_page_token")
        if not next_page_token:
            break

        pagina += 1

    return resenas


def generar_respuesta(resena: dict, nombre: str, tipo: str, tono: str) -> str:
    prompt = f"""Eres el community manager de {nombre}, un {tipo}.
Tu tono de comunicación es {tono}.

{resena['autor']} dejó esta reseña de Google con {resena['estrellas']} estrellas:
"{resena['texto']}"

ESTRUCTURA OBLIGATORIA:
1. Hook de primera línea: máximo 6 palabras, genera urgencia o curiosidad
2. Respuesta personalizada mencionando detalles específicos de la reseña
3. Una sola CTA al final

REGLAS:
- Máximo 80 palabras
- Si es positiva (4-5 estrellas): agradece específicamente, invita a volver
- Si es negativa (1-2 estrellas): empatía, reconoce el problema, ofrece resolverlo fuera de la plataforma
- Si es neutral (3 estrellas): agradece, reconoce el punto de mejora, invita a volver
- Sin hashtags, en español latino, tono directo y confiable
- Termina con el nombre del negocio

Devuelve SOLO la respuesta lista para publicar, sin explicaciones."""

    message = anthropic_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}]
    )
    return message.content[0].text.strip()


def guardar_excel(resultados: list[dict], nombre_negocio: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = nombre_negocio.replace(" ", "_").replace("/", "-")
    archivo = f"RESPUESTAS_{nombre_limpio}_{timestamp}.xlsx"
    ruta = os.path.join(os.path.dirname(__file__), archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas"

    # Colores
    verde = PatternFill("solid", fgColor="1F7A4A")
    gris_claro = PatternFill("solid", fgColor="F2F2F2")
    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Respuestas generadas para: {nombre_negocio}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = verde
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Total: {len(resultados)} reseñas"
    ws["A2"].font = Font(italic=True, size=10, color="888888")
    ws.merge_cells("A2:F2")

    # Encabezados
    encabezados = ["#", "Estrellas", "Autor", "Fecha", "Reseña del cliente", "✅ Respuesta (copiar y pegar en Google)"]
    anchos = [5, 10, 20, 15, 50, 60]

    for col, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
        celda = ws.cell(row=3, column=col, value=enc)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = verde
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.row_dimensions[3].height = 25

    # Datos
    for i, r in enumerate(resultados, 1):
        fila = i + 3
        fill = gris_claro if i % 2 == 0 else PatternFill()
        estrellas_texto = "⭐" * r["estrellas"] + f" ({r['estrellas']})"

        valores = [i, estrellas_texto, r["autor"], r["fecha"], r["resena_original"], r["respuesta_generada"]]

        for col, valor in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = borde
            if i % 2 == 0:
                celda.fill = gris_claro

        ws.row_dimensions[fila].height = 80

    # Congelar encabezados
    ws.freeze_panes = "A4"

    wb.save(ruta)
    return archivo


def guardar_resultados(resultados: list[dict], nombre_negocio: str, data_id: str):
    nombre_limpio = nombre_negocio.replace(" ", "_").replace("/", "-")

    archivo_excel = guardar_excel(resultados, nombre_negocio)

    # Guardar registro para el script de mantenimiento
    registro = {
        "nombre_negocio": nombre_negocio,
        "data_id": data_id,
        "tipo": resultados[0].get("tipo", "restaurante") if resultados else "",
        "tono": resultados[0].get("tono", "profesional y cercano") if resultados else "",
        "ultimo_proceso": datetime.now().isoformat(),
        "total_procesadas": len(resultados)
    }

    archivo_registro = f"registro_{nombre_limpio}.json"
    ruta_registro = os.path.join(os.path.dirname(__file__), archivo_registro)
    with open(ruta_registro, "w", encoding="utf-8") as f:
        json.dump(registro, f, ensure_ascii=False, indent=2)

    print(f"\nExcel guardado en:    {archivo_excel}")
    print(f"Registro guardado en: {archivo_registro} (úsalo para el mantenimiento)")


def main():
    print("\n=== SETUP COMPLETO — TODAS LAS RESEÑAS ===\n")

    serpapi_key = os.environ.get("SERPAPI_KEY") or input("SerpAPI key: ").strip()

    nombre_busqueda = input("Nombre y ciudad del negocio (ej: Brasas Grill Nashville): ").strip()
    tipo = input("Tipo de negocio [restaurante]: ").strip() or "restaurante"
    tono = input("Tono de respuesta [profesional y cercano]: ").strip() or "profesional y cercano"

    data_id, nombre_negocio = buscar_data_id(nombre_busqueda, serpapi_key)

    print(f"\nDescargando todas las reseñas de {nombre_negocio}...")
    resenas = obtener_todas_las_resenas(data_id, serpapi_key)

    if not resenas:
        print("No se encontraron reseñas.")
        return

    print(f"\nTotal de reseñas encontradas: {len(resenas)}")
    print(f"Generando respuestas...")

    resultados = []
    for i, r in enumerate(resenas, 1):
        print(f"  Procesando {i}/{len(resenas)}...", end=" ")
        respuesta = generar_respuesta(r, nombre_negocio, tipo, tono)
        resultados.append({
            "estrellas": r["estrellas"],
            "autor": r["autor"],
            "fecha": r["fecha"],
            "resena_original": r["texto"],
            "respuesta_generada": respuesta,
            "tipo": tipo,
            "tono": tono
        })
        print("✓")

    print("\n" + "=" * 60)
    for i, r in enumerate(resultados, 1):
        print(f"\nRESEÑA #{i} — {'⭐' * r['estrellas']} — {r['autor']}")
        print(f"Cliente: {r['resena_original']}")
        print(f"\nRespuesta:\n{r['respuesta_generada']}")
        print("-" * 60)

    guardar_resultados(resultados, nombre_negocio, data_id)


if __name__ == "__main__":
    main()
